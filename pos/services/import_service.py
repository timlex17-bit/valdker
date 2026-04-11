from io import BytesIO

from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from pos.models import (
    Category,
    Unit,
    Product,
    Customer,
    Supplier,
    StockMovement,
)
from pos.models_backup import BackupHistory
from pos.models_import import ImportJob, ImportRowError


# =========================================================
# TEMPLATE
# =========================================================
TEMPLATE_SHEETS = {
    "Categories": ["name"],
    "Units": ["name"],
    "Products": [
        "name",
        "code",
        "sku",
        "item_type",
        "category",
        "unit",
        "supplier",
        "buy_price",
        "sell_price",
        "stock",
        "track_stock",
        "description",
        "is_active",
    ],
    "Customers": ["name", "cell", "email", "address", "points"],
    "Suppliers": ["name", "contact_person", "cell", "email", "address"],
    "OpeningStock": ["product_code", "quantity"],
}


# =========================================================
# TEMPLATE BUILDERS
# =========================================================
def build_template_workbook():
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name, headers in TEMPLATE_SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def template_info():
    return {
        "filename": "ValdKerPOS_Master_Import_Template.xlsx",
        "format": "Excel Workbook",
        "sheets": list(TEMPLATE_SHEETS.keys()),
        "description": "One file with multiple sheets for importing master data.",
    }


# =========================================================
# FILE / VALUE HELPERS
# =========================================================
def _load_workbook(file_path: str):
    try:
        return load_workbook(file_path, data_only=True)
    except FileNotFoundError:
        raise ValueError("Import file not found.")
    except Exception as e:
        raise ValueError(f"Failed to read workbook: {e}")


def _safe_str(value):
    return str(value).strip() if value is not None else ""


def _safe_int(value, default=0):
    if value in ("", None):
        return default
    try:
        return int(float(value))
    except Exception:
        raise ValueError("Must be a numeric value.")


def _safe_decimal(value, default=0):
    if value in ("", None):
        return default
    try:
        return float(value)
    except Exception:
        raise ValueError("Must be a numeric value.")


def _safe_bool(value, default=False):
    if value in (True, False):
        return value
    if value in ("", None):
        return default

    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y"}:
        return True
    if text in {"0", "false", "no", "n"}:
        return False
    return default


def _sheet_headers(ws):
    return [_safe_str(c.value) for c in ws[1]] if ws.max_row >= 1 else []


def _row_to_dict(headers, row):
    data = {}
    row = row or []
    for idx, header in enumerate(headers):
        data[header] = row[idx] if idx < len(row) else None
    return data


def _clear_previous_errors(import_job: ImportJob):
    import_job.row_errors.all().delete()


def _add_error(import_job: ImportJob, sheet_name: str, row_number: int, field_name: str, message: str):
    ImportRowError.objects.create(
        import_job=import_job,
        sheet_name=sheet_name,
        row_number=row_number,
        field_name=field_name,
        message=message,
    )


def _is_blank_row(row_values):
    if not row_values:
        return True
    return not any(v not in (None, "", " ") for v in row_values)


def _normalize_sheet_name(sheet_name: str):
    return _safe_str(sheet_name)


# =========================================================
# VALIDATION LOGIC PER SHEET
# =========================================================
def _validate_headers(sheet_name: str, headers: list[str]):
    expected = TEMPLATE_SHEETS.get(sheet_name)
    if expected is None:
        return [{
            "field_name": "__sheet__",
            "message": f"Unknown sheet '{sheet_name}'.",
        }]

    if headers != expected:
        return [{
            "field_name": "__header__",
            "message": f"Invalid header for sheet '{sheet_name}'. Expected: {expected}",
        }]

    return []


def _validate_row_by_sheet(
    sheet_name: str,
    row_data: dict,
    *,
    existing_categories=None,
    existing_units=None,
    existing_suppliers=None,
    existing_products=None,
    file_categories=None,
    file_units=None,
    file_suppliers=None,
    file_products=None,
):
    errors = []

    existing_categories = existing_categories or set()
    existing_units = existing_units or set()
    existing_suppliers = existing_suppliers or set()
    existing_products = existing_products or set()

    file_categories = file_categories or set()
    file_units = file_units or set()
    file_suppliers = file_suppliers or set()
    file_products = file_products or set()

    if sheet_name == "Categories":
        name = _safe_str(row_data.get("name"))
        if not name:
            errors.append({"field_name": "name", "message": "Category name is required."})

    elif sheet_name == "Units":
        name = _safe_str(row_data.get("name"))
        if not name:
            errors.append({"field_name": "name", "message": "Unit name is required."})

    elif sheet_name == "Customers":
        name = _safe_str(row_data.get("name"))
        if not name:
            errors.append({"field_name": "name", "message": "Customer name is required."})

        try:
            _safe_int(row_data.get("points"), default=0)
        except ValueError as e:
            errors.append({"field_name": "points", "message": str(e)})

    elif sheet_name == "Suppliers":
        name = _safe_str(row_data.get("name"))
        if not name:
            errors.append({"field_name": "name", "message": "Supplier name is required."})

    elif sheet_name == "Products":
        name = _safe_str(row_data.get("name"))
        code = _safe_str(row_data.get("code"))
        item_type = _safe_str(row_data.get("item_type")).lower()
        category_name = _safe_str(row_data.get("category"))
        unit_name = _safe_str(row_data.get("unit"))
        supplier_name = _safe_str(row_data.get("supplier"))

        if not name:
            errors.append({"field_name": "name", "message": "Product name is required."})

        if not code:
            errors.append({"field_name": "code", "message": "Product code is required."})

        if item_type and item_type not in {"product", "service", "part"}:
            errors.append({
                "field_name": "item_type",
                "message": "Item type must be one of: product, service, part.",
            })

        try:
            _safe_decimal(row_data.get("buy_price"), default=0)
        except ValueError as e:
            errors.append({"field_name": "buy_price", "message": str(e)})

        try:
            _safe_decimal(row_data.get("sell_price"), default=0)
        except ValueError as e:
            errors.append({"field_name": "sell_price", "message": str(e)})

        try:
            _safe_int(row_data.get("stock"), default=0)
        except ValueError as e:
            errors.append({"field_name": "stock", "message": str(e)})

        if category_name:
            category_key = category_name.lower()
            if category_key not in existing_categories and category_key not in file_categories:
                errors.append({
                    "field_name": "category",
                    "message": f"Category '{category_name}' not found in this file or this shop.",
                })

        if unit_name:
            unit_key = unit_name.lower()
            if unit_key not in existing_units and unit_key not in file_units:
                errors.append({
                    "field_name": "unit",
                    "message": f"Unit '{unit_name}' not found in this file or this shop.",
                })

        if supplier_name:
            supplier_key = supplier_name.lower()
            if supplier_key not in existing_suppliers and supplier_key not in file_suppliers:
                errors.append({
                    "field_name": "supplier",
                    "message": f"Supplier '{supplier_name}' not found in this file or this shop.",
                })

    elif sheet_name == "OpeningStock":
        product_code = _safe_str(row_data.get("product_code"))
        quantity = row_data.get("quantity")

        if not product_code:
            errors.append({"field_name": "product_code", "message": "Product code is required."})
        else:
            product_key = product_code.lower()
            if product_key not in existing_products and product_key not in file_products:
                errors.append({
                    "field_name": "product_code",
                    "message": f"Product with code '{product_code}' not found in this file or this shop.",
                })

        try:
            _safe_int(quantity, default=0)
        except ValueError as e:
            errors.append({"field_name": "quantity", "message": str(e)})

    else:
        errors.append({
            "field_name": "__sheet__",
            "message": f"Unsupported sheet '{sheet_name}'.",
        })

    return errors


# =========================================================
# VALIDATION
# =========================================================
def validate_import_workbook(import_job: ImportJob):
    workbook = _load_workbook(import_job.file.path)
    _clear_previous_errors(import_job)

    shop = import_job.shop

    total_rows = 0
    valid_rows = 0
    invalid_rows = 0

    preview_data = {
        "Categories": [],
        "Units": [],
        "Products": [],
        "Customers": [],
        "Suppliers": [],
        "OpeningStock": [],
    }

    # preload existing data from DB once
    existing_categories = set(
        Category.objects.filter(shop=shop).values_list("name", flat=True)
    )
    existing_units = set(
        Unit.objects.filter(shop=shop).values_list("name", flat=True)
    )
    existing_suppliers = set(
        Supplier.objects.filter(shop=shop).values_list("name", flat=True)
    )
    existing_products = set(
        Product.objects.filter(shop=shop).values_list("code", flat=True)
    )

    existing_categories = {str(x).strip().lower() for x in existing_categories if x}
    existing_units = {str(x).strip().lower() for x in existing_units if x}
    existing_suppliers = {str(x).strip().lower() for x in existing_suppliers if x}
    existing_products = {str(x).strip().lower() for x in existing_products if x}

    # cache references found inside the uploaded workbook
    file_categories = set()
    file_units = set()
    file_suppliers = set()
    file_products = set()

    header_has_error = False

    # -----------------------------------------------------
    # PASS 1: collect in-file references
    # -----------------------------------------------------
    for sheet_name in workbook.sheetnames:
        normalized_sheet_name = _normalize_sheet_name(sheet_name)
        if normalized_sheet_name not in TEMPLATE_SHEETS:
            continue

        ws = workbook[sheet_name]
        headers = _sheet_headers(ws)

        if headers != TEMPLATE_SHEETS[normalized_sheet_name]:
            continue

        if ws.max_row < 2:
            continue

        for _, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_values = list(row or [])
            if _is_blank_row(row_values):
                continue

            row_data = _row_to_dict(headers, row_values)

            if normalized_sheet_name == "Categories":
                name = _safe_str(row_data.get("name"))
                if name:
                    file_categories.add(name.lower())

            elif normalized_sheet_name == "Units":
                name = _safe_str(row_data.get("name"))
                if name:
                    file_units.add(name.lower())

            elif normalized_sheet_name == "Suppliers":
                name = _safe_str(row_data.get("name"))
                if name:
                    file_suppliers.add(name.lower())

            elif normalized_sheet_name == "Products":
                code = _safe_str(row_data.get("code"))
                if code:
                    file_products.add(code.lower())

    # -----------------------------------------------------
    # PASS 2: validate rows
    # -----------------------------------------------------
    for sheet_name in workbook.sheetnames:
        normalized_sheet_name = _normalize_sheet_name(sheet_name)
        ws = workbook[sheet_name]
        headers = _sheet_headers(ws)

        header_errors = _validate_headers(normalized_sheet_name, headers)
        if header_errors:
            header_has_error = True
            for err in header_errors:
                _add_error(
                    import_job=import_job,
                    sheet_name=normalized_sheet_name,
                    row_number=1,
                    field_name=err.get("field_name", "__header__"),
                    message=err.get("message", "Invalid header."),
                )
            continue

        if ws.max_row < 2:
            continue

        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_values = list(row or [])
            if _is_blank_row(row_values):
                continue

            excel_row_number = row_index
            total_rows += 1

            row_data = _row_to_dict(headers, row_values)

            if len(preview_data.get(normalized_sheet_name, [])) < 20:
                safe_row = {}
                for key, value in row_data.items():
                    if value is None:
                        safe_row[key] = ""
                    elif isinstance(value, (str, int, float, bool)):
                        safe_row[key] = value
                    else:
                        safe_row[key] = str(value)

                preview_data.setdefault(normalized_sheet_name, []).append({
                    "row_number": excel_row_number,
                    "data": safe_row,
                })

            try:
                row_errors = _validate_row_by_sheet(
                    sheet_name=normalized_sheet_name,
                    row_data=row_data,
                    existing_categories=existing_categories,
                    existing_units=existing_units,
                    existing_suppliers=existing_suppliers,
                    existing_products=existing_products,
                    file_categories=file_categories,
                    file_units=file_units,
                    file_suppliers=file_suppliers,
                    file_products=file_products,
                )
            except Exception as e:
                row_errors = [{
                    "field_name": "__all__",
                    "message": str(e),
                }]

            if row_errors:
                invalid_rows += 1
                for err in row_errors:
                    _add_error(
                        import_job=import_job,
                        sheet_name=normalized_sheet_name,
                        row_number=excel_row_number,
                        field_name=err.get("field_name", "__all__"),
                        message=err.get("message", "Invalid row."),
                    )
            else:
                valid_rows += 1

    has_errors = header_has_error or invalid_rows > 0

    import_job.status = (
        ImportJob.Status.VALIDATED
        if not has_errors
        else ImportJob.Status.UPLOADED
    )
    import_job.metadata = {
        **(import_job.metadata or {}),
        "validation_summary": {
            "total_rows": total_rows,
            "valid_rows": valid_rows,
            "invalid_rows": invalid_rows,
            "has_errors": has_errors,
        },
        "preview": preview_data,
        "validated_at": timezone.localtime().isoformat(),
    }
    import_job.save(update_fields=["status", "metadata"])

    return {
        "import_job_id": import_job.id,
        "status": import_job.status,
        "total_rows": total_rows,
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "errors": [
            {
                "sheet_name": err.sheet_name,
                "row_number": err.row_number,
                "field_name": err.field_name,
                "message": err.message,
            }
            for err in import_job.row_errors.all().order_by("sheet_name", "row_number", "id")
        ],
        "preview": preview_data,
    }

# =========================================================
# IMPORT EXECUTION
# =========================================================
def _latest_successful_backup_exists(shop, hours=24):
    cutoff = timezone.now() - timezone.timedelta(hours=hours)
    return BackupHistory.objects.filter(
        shop=shop,
        status=BackupHistory.Status.SUCCESS,
        deleted_at__isnull=True,
        completed_at__gte=cutoff,
    ).exists()


@transaction.atomic
def run_import(import_job: ImportJob, *, confirm_import: bool, skip_backup_check: bool = False):
    if not confirm_import:
        raise ValueError("Import confirmation is required.")

    if import_job.status not in {ImportJob.Status.VALIDATED, ImportJob.Status.UPLOADED}:
        raise ValueError("Import job must be uploaded or validated before import.")

    if import_job.status == ImportJob.Status.UPLOADED:
        validate_import_workbook(import_job)

    current_invalid_rows = import_job.row_errors.count()
    if current_invalid_rows > 0:
        raise ValueError("Import cannot continue because invalid rows still exist.")

    if not skip_backup_check and not _latest_successful_backup_exists(import_job.shop):
        raise ValueError("No recent successful backup found. Please create a backup before import.")

    import_job.mark_importing()

    wb = load_workbook(import_job.file.path, data_only=True)
    shop = import_job.shop

    imported_rows = 0
    skipped_rows = 0

    # =====================================================
    # PRELOAD EXISTING DATA
    # =====================================================
    existing_categories = {
        obj.name.strip().lower(): obj
        for obj in Category.objects.filter(shop=shop)
        if obj.name
    }
    existing_units = {
        obj.name.strip().lower(): obj
        for obj in Unit.objects.filter(shop=shop)
        if obj.name
    }
    existing_suppliers = {
        obj.name.strip().lower(): obj
        for obj in Supplier.objects.filter(shop=shop)
        if obj.name
    }
    existing_customers = {
        obj.name.strip().lower(): obj
        for obj in Customer.objects.filter(shop=shop)
        if obj.name
    }
    existing_products = {
        obj.code.strip().lower(): obj
        for obj in Product.objects.filter(shop=shop)
        if obj.code
    }

    # =====================================================
    # CATEGORIES - BULK CREATE
    # =====================================================
    category_to_create = []
    if "Categories" in wb.sheetnames:
        ws = wb["Categories"]
        headers = _sheet_headers(ws)

        if headers == TEMPLATE_SHEETS["Categories"]:
            seen_in_file = set()

            for row in ws.iter_rows(min_row=2, values_only=True):
                if _is_blank_row(row):
                    continue

                data = _row_to_dict(headers, row)
                name = _safe_str(data.get("name"))
                if not name:
                    skipped_rows += 1
                    continue

                key = name.lower()
                if key in existing_categories or key in seen_in_file:
                    skipped_rows += 1
                    continue

                seen_in_file.add(key)
                category_to_create.append(Category(shop=shop, name=name))

            if category_to_create:
                Category.objects.bulk_create(category_to_create, batch_size=1000)
                imported_rows += len(category_to_create)

            existing_categories = {
                obj.name.strip().lower(): obj
                for obj in Category.objects.filter(shop=shop)
                if obj.name
            }

    # =====================================================
    # UNITS - BULK CREATE
    # =====================================================
    unit_to_create = []
    if "Units" in wb.sheetnames:
        ws = wb["Units"]
        headers = _sheet_headers(ws)

        if headers == TEMPLATE_SHEETS["Units"]:
            seen_in_file = set()

            for row in ws.iter_rows(min_row=2, values_only=True):
                if _is_blank_row(row):
                    continue

                data = _row_to_dict(headers, row)
                name = _safe_str(data.get("name"))
                if not name:
                    skipped_rows += 1
                    continue

                key = name.lower()
                if key in existing_units or key in seen_in_file:
                    skipped_rows += 1
                    continue

                seen_in_file.add(key)
                unit_to_create.append(Unit(shop=shop, name=name))

            if unit_to_create:
                Unit.objects.bulk_create(unit_to_create, batch_size=1000)
                imported_rows += len(unit_to_create)

            existing_units = {
                obj.name.strip().lower(): obj
                for obj in Unit.objects.filter(shop=shop)
                if obj.name
            }

    # =====================================================
    # SUPPLIERS - BULK CREATE
    # =====================================================
    supplier_to_create = []
    if "Suppliers" in wb.sheetnames:
        ws = wb["Suppliers"]
        headers = _sheet_headers(ws)

        if headers == TEMPLATE_SHEETS["Suppliers"]:
            seen_in_file = set()

            for row in ws.iter_rows(min_row=2, values_only=True):
                if _is_blank_row(row):
                    continue

                data = _row_to_dict(headers, row)
                name = _safe_str(data.get("name"))
                if not name:
                    skipped_rows += 1
                    continue

                key = name.lower()
                if key in existing_suppliers or key in seen_in_file:
                    skipped_rows += 1
                    continue

                seen_in_file.add(key)
                supplier_to_create.append(
                    Supplier(
                        shop=shop,
                        name=name,
                        contact_person=_safe_str(data.get("contact_person")),
                        cell=_safe_str(data.get("cell")),
                        email=_safe_str(data.get("email")) or None,
                        address=_safe_str(data.get("address")),
                    )
                )

            if supplier_to_create:
                Supplier.objects.bulk_create(supplier_to_create, batch_size=1000)
                imported_rows += len(supplier_to_create)

            existing_suppliers = {
                obj.name.strip().lower(): obj
                for obj in Supplier.objects.filter(shop=shop)
                if obj.name
            }

    # =====================================================
    # CUSTOMERS - BULK CREATE
    # =====================================================
    customer_to_create = []
    if "Customers" in wb.sheetnames:
        ws = wb["Customers"]
        headers = _sheet_headers(ws)

        if headers == TEMPLATE_SHEETS["Customers"]:
            seen_in_file = set()

            for row in ws.iter_rows(min_row=2, values_only=True):
                if _is_blank_row(row):
                    continue

                data = _row_to_dict(headers, row)
                name = _safe_str(data.get("name"))
                if not name:
                    skipped_rows += 1
                    continue

                key = name.lower()
                if key in existing_customers or key in seen_in_file:
                    skipped_rows += 1
                    continue

                seen_in_file.add(key)
                customer_to_create.append(
                    Customer(
                        shop=shop,
                        name=name,
                        cell=_safe_str(data.get("cell")),
                        email=_safe_str(data.get("email")) or None,
                        address=_safe_str(data.get("address")),
                        points=_safe_int(data.get("points"), default=0),
                    )
                )

            if customer_to_create:
                Customer.objects.bulk_create(customer_to_create, batch_size=1000)
                imported_rows += len(customer_to_create)

            existing_customers = {
                obj.name.strip().lower(): obj
                for obj in Customer.objects.filter(shop=shop)
                if obj.name
            }

    # refresh references after master import
    category_map = existing_categories
    unit_map = existing_units
    supplier_map = existing_suppliers
    product_map = existing_products.copy()

    # =====================================================
    # PRODUCTS - BULK CREATE + BULK UPDATE
    # =====================================================
    products_to_create = []
    products_to_update = []
    seen_product_codes = set()

    if "Products" in wb.sheetnames:
        ws = wb["Products"]
        headers = _sheet_headers(ws)

        if headers == TEMPLATE_SHEETS["Products"]:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if _is_blank_row(row):
                    continue

                data = _row_to_dict(headers, row)

                code = _safe_str(data.get("code"))
                name = _safe_str(data.get("name"))
                if not code or not name:
                    skipped_rows += 1
                    continue

                code_key = code.lower()

                # hindari duplikat code dalam file yang sama
                if code_key in seen_product_codes:
                    skipped_rows += 1
                    continue
                seen_product_codes.add(code_key)

                category_name = _safe_str(data.get("category")).lower()
                unit_name = _safe_str(data.get("unit")).lower()
                supplier_name = _safe_str(data.get("supplier")).lower()

                category = category_map.get(category_name) if category_name else None
                unit = unit_map.get(unit_name) if unit_name else None
                supplier = supplier_map.get(supplier_name) if supplier_name else None

                prepared = {
                    "name": name,
                    "sku": _safe_str(data.get("sku")) or None,
                    "item_type": _safe_str(data.get("item_type")).lower() or "product",
                    "category": category,
                    "unit": unit,
                    "supplier": supplier,
                    "description": _safe_str(data.get("description")),
                    "stock": _safe_int(data.get("stock"), default=0),
                    "track_stock": _safe_bool(data.get("track_stock"), default=True),
                    "buy_price": _safe_decimal(data.get("buy_price"), default=0),
                    "sell_price": _safe_decimal(data.get("sell_price"), default=0),
                    "weight": 0,
                    "is_active": _safe_bool(data.get("is_active"), default=True),
                }

                existing = product_map.get(code_key)
                if existing:
                    existing.name = prepared["name"]
                    existing.sku = prepared["sku"]
                    existing.item_type = prepared["item_type"]
                    existing.category = prepared["category"]
                    existing.unit = prepared["unit"]
                    existing.supplier = prepared["supplier"]
                    existing.description = prepared["description"]
                    existing.track_stock = prepared["track_stock"]
                    existing.buy_price = prepared["buy_price"]
                    existing.sell_price = prepared["sell_price"]
                    existing.is_active = prepared["is_active"]
                    products_to_update.append(existing)
                else:
                    products_to_create.append(
                        Product(
                            shop=shop,
                            code=code,
                            name=prepared["name"],
                            sku=prepared["sku"],
                            item_type=prepared["item_type"],
                            category=prepared["category"],
                            unit=prepared["unit"],
                            supplier=prepared["supplier"],
                            description=prepared["description"],
                            stock=prepared["stock"],
                            track_stock=prepared["track_stock"],
                            buy_price=prepared["buy_price"],
                            sell_price=prepared["sell_price"],
                            weight=prepared["weight"],
                            is_active=prepared["is_active"],
                        )
                    )

            if products_to_create:
                Product.objects.bulk_create(products_to_create, batch_size=1000)
                imported_rows += len(products_to_create)

            if products_to_update:
                Product.objects.bulk_update(
                    products_to_update,
                    fields=[
                        "name",
                        "sku",
                        "item_type",
                        "category",
                        "unit",
                        "supplier",
                        "description",
                        "track_stock",
                        "buy_price",
                        "sell_price",
                        "is_active",
                    ],
                    batch_size=1000,
                )
                imported_rows += len(products_to_update)

            product_map = {
                obj.code.strip().lower(): obj
                for obj in Product.objects.filter(shop=shop)
                if obj.code
            }

    # =====================================================
    # OPENING STOCK - BULK UPDATE + BULK MOVEMENTS
    # =====================================================
    stock_products_to_update = []
    stock_movements_to_create = []
    touched_product_ids = set()

    if "OpeningStock" in wb.sheetnames:
        ws = wb["OpeningStock"]
        headers = _sheet_headers(ws)

        if headers == TEMPLATE_SHEETS["OpeningStock"]:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if _is_blank_row(row):
                    continue

                data = _row_to_dict(headers, row)
                product_code = _safe_str(data.get("product_code")).lower()
                qty = _safe_int(data.get("quantity"), default=0)

                if not product_code:
                    skipped_rows += 1
                    continue

                product = product_map.get(product_code)
                if not product:
                    skipped_rows += 1
                    continue

                before_stock = product.stock
                after_stock = qty

                if before_stock == after_stock:
                    skipped_rows += 1
                    continue

                # hindari product yang sama diupdate berulang di file opening stock
                if product.id in touched_product_ids:
                    skipped_rows += 1
                    continue

                touched_product_ids.add(product.id)
                product.stock = after_stock
                stock_products_to_update.append(product)

                stock_movements_to_create.append(
                    StockMovement(
                        shop=shop,
                        product=product,
                        movement_type=StockMovement.Type.ADJUSTMENT,
                        quantity_delta=after_stock - before_stock,
                        before_stock=before_stock,
                        after_stock=after_stock,
                        note=f"Opening stock import #{import_job.id}",
                        ref_model="ImportJob",
                        ref_id=import_job.id,
                        created_by=import_job.uploaded_by,
                    )
                )
                imported_rows += 1

            if stock_products_to_update:
                Product.objects.bulk_update(
                    stock_products_to_update,
                    fields=["stock"],
                    batch_size=1000,
                )

            if stock_movements_to_create:
                StockMovement.objects.bulk_create(stock_movements_to_create, batch_size=1000)

    import_job.mark_completed(
        imported_rows=imported_rows,
        skipped_rows=skipped_rows,
        note="Master data import completed successfully.",
        metadata={
            **(import_job.metadata or {}),
            "imported_at": timezone.localtime().isoformat(),
        },
    )

    return import_job